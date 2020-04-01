import pprint
import os
import pandas as pd
import openpyxl
import xlrd
import numpy as np
from openpyxl import load_workbook

#change directory
os.chdir('/Users/jiwoonhwang/Desktop/MBTS/')
path ="/Users/jiwoonhwang/Desktop/MBTS/Pooled_pos_bethanie.xlsx"


#wb = openpyxl.load_workbook(path)
#sheet = wb.active
#sheet_obj = wb_obj.get_sheet_by_name("")
#sheet_obj = wb_obj.worksheets[#]
#max_col = sheet_obj.max_column

#sheet_obj.insert_cols(2)
#new = sheet_obj.cell(row=1, column=2)


#for i in range(1, max_col+1):
#	cell_obj = sheet_obj.cell(row=1, column=i)
#	print(cell_obj.value)


df =pd.read_excel(path, sheet_name='Pooled_pos_con')

#REPLACE ZERO VALUES WITH NAN
df1=df.replace(0,np.NaN)

df2 = pd.DataFrame()
df2 = df1.iloc[:,0:12]


depth = []
for col in df.columns:
		column_split=col.split('_')
		if len(column_split)==4:
			if column_split[0]=='MARS':
				if column_split[2] not in depth:
					depth.append(column_split[2])
					#print depth
cast_= []
for col in df.columns:
		column_split=col.split('_')
		if len(column_split)==4:
			if column_split[0]=='MARS':
				if column_split[1] not in cast_:
					cast_.append(column_split[1])
					#print cast_

depth = "0m"	
for cst in cast_:
	countlist = []
	count = 0
	for col in df.columns:
		count += 1
		column_split=col.split('_')	
		if len(column_split)==4:
			if column_split[1]==cst:
				if column_split[2]==depth:								
					countlist.append(count)
					cast = column_split[1]
	print countlist
	print cast
					#print cast
				#print dep
	if len(countlist)!=0:			
		start = countlist[0]
		end = countlist[-1]	
		choose = df1.iloc[:,start-1:end]
		choose_aver = choose.mean(axis=1)

		#df1.insert(int(end),"average_"+cast+"_"+depth,choose_aver, True)

		x= "average_"+cast+"_"+depth
		df2[x]= choose_aver

#print "End of depth 0m"

depth = "20m"	
for cst in cast_:
	countlist = []
	count = 0
	for col in df.columns:
		count += 1
		column_split=col.split('_')	
		if len(column_split)>=4:
			if column_split[1]==cst:
				if column_split[2]==depth:								
					countlist.append(count)
					cast = column_split[1]
	print countlist
	print cast
					#print cast
				#print dep
	if len(countlist)!=0:			
		start = countlist[0]
		end = countlist[-1]	
		choose = df1.iloc[:,start-1:end]
		choose_aver = choose.mean(axis=1)

		#df1.insert(int(end),"average_"+cast+"_"+depth,choose_aver, True)

		x= "average_"+cast+"_"+depth
		df2[x]= choose_aver

df1 = df1.fillna(0)
df1.to_excel('LOB_Peaklist_neg_edited_1.xlsx')

#CREATE NEW WORKSHEET WITH AVERAGES
writer = pd.ExcelWriter('LOB_Peaklist_neg_edited_1.xlsx', engine='openpyxl')
if os.path.exists('LOB_Peaklist_neg_edited_1.xlsx'):
	book = load_workbook('LOB_Peaklist_neg_edited_1.xlsx')
	writer.book = book

df2=df2.fillna(0)
df2.to_excel(writer, sheet_name='average')
writer.save()
writer.close()