import pprint
import os
import pandas as pd
import openpyxl
import xlrd
import numpy as np
from openpyxl import load_workbook

#change directory
os.chdir('/Users/jiwoonhwang/Desktop/MBTS/C2/')
path ="/Users/jiwoonhwang/Desktop/MBTS/pooled_pos_worksheet.xlsx"


#wb = openpyxl.load_workbook(path)
#sheet = wb.active
#sheet_obj = wb_obj.get_sheet_by_name("")
#sheet_obj = wb_obj.worksheets[#]
#max_col = sheet_obj.max_column

#sheet_obj.insert_cols(2)
#new = sheet_obj.cell(row=1, column=2)


#for i in range(1, max_col+1):
#	cell_obj = sheet_obj.cell(row=1, column=i)
#	print(cell_obj.value)


df =pd.read_excel(path, sheet_name='we_diel')

#REPLACE ZERO VALUES WITH NAN
df1=df.replace(0,np.NaN)

df2 = pd.DataFrame()
df2 = df1.iloc[:,0:12]


depth = []
for col in df.columns:
		column_split=col.split('_')
		if len(column_split)==4:
			if column_split[0]=='MARS':
				if column_split[2] not in depth:
					depth.append(column_split[2])
					#print depth
cast_= []
for col in df.columns:
		column_split=col.split('_')
		if len(column_split)==4:
			if column_split[0]=='MARS':
				if column_split[1] not in cast_:
					cast_.append(column_split[1])

for dep in reversed(depth):
    countlist = []
    count = 0
    for col in df.columns:
        count += 1
        columnmn_split=col.split('_')
        if len(column_split)==4:
            if column_split[2]== dep:
                countlist.append(count)
                cast = column_split[1]
                print countlist
    start = countlist[0]
    end = countlist[-1]
    choose = df1.iloc[:,start-1:end]
    choose_aver = choose.mean(axis=1)
	#print cast	
			
			
##average of entire row## 
#num_aver = df1.mean(axis=1)
#print num_aver

			#print (choose)
			#print (choose_aver)
	df1.insert(int(end),"average_"+cast+"_"+dep,choose_aver, True)

	x= "average_"+cast+"_"+dep
	df2[x]= choose_aver

df1 = df1.fillna(0)
df1.to_excel('LOB_Peaklist_neg_edited_1.xlsx')

#CREATE NEW WORKSHEET WITH AVERAGES
writer = pd.ExcelWriter('LOB_Peaklist_neg_edited_1.xlsx', engine='openpyxl')
if os.path.exists('LOB_Peaklist_neg_edited_1.xlsx'):
	book = load_workbook('LOB_Peaklist_neg_edited_1.xlsx')
	writer.book = book

df2=df2.fillna(0)
df2.to_excel(writer, sheet_name='average')
writer.save()
writer.close()	
#print df2

#for col in df1.columns:
#	print col





#open csv file
#wb = open('LOB_Peaklist_neg.csv')
 
#csvreader = csv.reader(wb)

#column header
#fields =[]
#fields = csvreader.next()
#print fields

#row by row
#rows = []
#for row in csvreader:
	#rows.append(row)
	#print row[1]


#for row in rows[:3]:
	#for col in row:
		#print ("%10s"%col),
	#print('\n')



		
